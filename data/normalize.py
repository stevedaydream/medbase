import pandas as pd
import re
import openpyxl

def fix_enc(v):
    """修正 pandas 從 Excel 讀到的 mojibake（Big5 bytes 被誤讀為 latin-1）。"""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return v
    s = str(v)
    try:
        return s.encode('latin-1').decode('big5')
    except (UnicodeEncodeError, UnicodeDecodeError):
        return s

def fix_df(df):
    """對 DataFrame 所有 object 欄位套用 fix_enc。"""
    for col in df.select_dtypes(include='object').columns:
        df[col] = df[col].apply(lambda x: fix_enc(x) if isinstance(x, str) else x)
    return df

def parse_price(raw):
    """Return (price_int_or_None, extra_note_or_None) from a potentially messy price cell."""
    if pd.isna(raw):
        return None, None
    s = str(raw).strip()
    # purely numeric (int or float)
    try:
        return int(float(s)), None
    except ValueError:
        pass
    # leading digits  e.g. "121000    (共15萬七千)" or "5600(顧問用)"
    m = re.match(r'^(\d+)', s)
    if m:
        return int(m.group(1)), s[m.end():].strip() or None
    # no leading number
    return None, s or None
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path

# ─────────────────────────────────────────────
# READ
# ─────────────────────────────────────────────
xl = pd.read_excel('data/items_parsed.xlsx', sheet_name=None)
df_raw_items = fix_df(xl['自費單品'])
df_raw_sets  = fix_df(xl['套組'])

# ─────────────────────────────────────────────
# TABLE: items  (deduplicated by hospital_code)
# ─────────────────────────────────────────────
df_raw_items['price'] = pd.to_numeric(df_raw_items['price'], errors='coerce')
df_raw_items = df_raw_items.drop(columns=['Unnamed: 6'])

def pick_canonical(group):
    # group does NOT include the grouping key (hospital_code) due to include_groups=False
    no_star = group[~group['nameZH'].fillna('').str.startswith('★')]
    cands   = no_star if len(no_star) > 0 else group
    return cands.loc[cands.notna().sum(axis=1).idxmax()]

items_df = (
    df_raw_items
    .groupby('hospital_code', group_keys=False)
    .apply(pick_canonical, include_groups=False)
    .reset_index()          # hospital_code comes back as index → column
    .rename(columns={'nameEN': 'name_en', 'nameZH': 'name_zh', 'items': 'body_part'})
)
items_df = items_df[['hospital_code','name_en','name_zh','category','body_part','unit','price','supplier','notes']]
items_df['price'] = items_df['price'].astype('Int64')

# ─────────────────────────────────────────────
# TABLE: doctors  (from VS column in 套組)
# ─────────────────────────────────────────────
unique_vs = df_raw_sets['VS'].dropna().unique()
doctors_df = pd.DataFrame({
    'id':            range(1, len(unique_vs) + 1),
    'name':          unique_vs,
    'department_id': pd.array([pd.NA] * len(unique_vs), dtype='Int64'),
})
vs_to_id = dict(zip(doctors_df['name'], doctors_df['id']))

# ─────────────────────────────────────────────
# TABLE: departments  (empty – user-populated)
# ─────────────────────────────────────────────
departments_df = pd.DataFrame({'id': pd.Series(dtype='Int64'), 'name': pd.Series(dtype=str)})

# ─────────────────────────────────────────────
# TABLES: sets + set_items  (parse 套組 grouping)
# VS non-null row = new set header (also the first item of that set)
# All-NaN rows are separators – skip
# ─────────────────────────────────────────────
sets_list      = []
set_items_list = []
current_set_id = 0
set_item_id    = 0

for _, row in df_raw_sets.iterrows():
    if row.isna().all():
        continue

    if pd.notna(row['VS']):
        current_set_id += 1
        surgery = str(row['術式']).strip() if pd.notna(row['術式']) else ''
        vs_name = str(row['VS']).strip()
        sets_list.append({
            'id':            current_set_id,
            'name':          f"{vs_name} - {surgery}" if surgery else vs_name,
            'surgery_type':  surgery or None,
            'doctor_id':     vs_to_id.get(row['VS']),
            'department_id': None,
            'notes':         None,
        })

    # add item row (must have a 代碼)
    if current_set_id > 0 and pd.notna(row.get('代碼')):
        set_item_id += 1
        name_zh = str(row['中文名稱']).replace('\n', ' ').strip() if pd.notna(row['中文名稱']) else None
        raw_qty = row['數量']
        if pd.notna(raw_qty):
            try:
                qty = int(float(raw_qty))
            except (ValueError, TypeError):
                qty = None  # e.g. 'PRN'
        else:
            qty = 1
        qty_note = str(raw_qty) if qty is None else None
        base_note = str(row['備註']).strip() if pd.notna(row['備註']) else None
        combined_note = ' | '.join(filter(None, [qty_note, base_note])) or None
        set_items_list.append({
            'id':            set_item_id,
            'set_id':        current_set_id,
            'hospital_code': str(row['代碼']).strip().upper() if pd.notna(row['代碼']) else None,
            'name_en':       str(row['英文名稱']).strip() if pd.notna(row['英文名稱']) else None,
            'name_zh':       name_zh,
            'quantity':      qty,
            'price':         parse_price(row['醫院收費'])[0],
            'notes':         ' | '.join(filter(None, [parse_price(row['醫院收費'])[1], combined_note])) or None,
        })

sets_df = pd.DataFrame(sets_list)
sets_df['doctor_id']     = sets_df['doctor_id'].astype('Int64')
sets_df['department_id'] = pd.array([pd.NA] * len(sets_df), dtype='Int64')

set_items_df = pd.DataFrame(set_items_list)
set_items_df['price'] = pd.array(
    [x if x is not None else pd.NA for x in set_items_df['price']], dtype='Int64'
)

print(f"items      : {len(items_df)} rows  (raw: {len(df_raw_items)})")
print(f"doctors    : {len(doctors_df)} rows")
print(f"departments: {len(departments_df)} rows")
print(f"sets       : {len(sets_df)} rows")
print(f"set_items  : {len(set_items_df)} rows")

# ═════════════════════════════════════════════
# OUTPUT 1 – SQL
# ═════════════════════════════════════════════
def sql_val(v):
    if v is None or (hasattr(v, '__class__') and v.__class__.__name__ == 'NAType'):
        return 'NULL'
    try:
        if pd.isna(v):
            return 'NULL'
    except Exception:
        pass
    if isinstance(v, (int, float)):
        return str(int(v))
    s = str(v).replace("'", "''").replace('\n', ' ').replace('\r', '')
    return "'" + s + "'"

SCHEMA = """\
-- MedBase Normalized Schema
-- Auto-generated from items_parsed.xlsx
-- Tables: departments, doctors, items, sets, set_items

PRAGMA foreign_keys = ON;

-- ── departments ──────────────────────────────
CREATE TABLE IF NOT EXISTS departments (
    id   INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT    NOT NULL UNIQUE
);

-- ── doctors ──────────────────────────────────
CREATE TABLE IF NOT EXISTS doctors (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    name          TEXT    NOT NULL,
    department_id INTEGER REFERENCES departments(id)
);

-- ── items ────────────────────────────────────
CREATE TABLE IF NOT EXISTS items (
    hospital_code TEXT    PRIMARY KEY,
    name_en       TEXT,
    name_zh       TEXT,
    category      TEXT,
    body_part     TEXT,
    unit          TEXT,
    price         INTEGER,
    supplier      TEXT,
    notes         TEXT
);

-- ── sets ─────────────────────────────────────
-- set_type 由以下欄位推斷：
--   doctor_id=有, department_id=無  → 主治醫師個人常用
--   doctor_id=無, department_id=有  → 科別標準套組
--   doctor_id=有, department_id=有  → 科別下特定醫師常用
--   兩者皆無                        → 依術式通用套組
CREATE TABLE IF NOT EXISTS sets (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    name          TEXT    NOT NULL,
    surgery_type  TEXT,
    doctor_id     INTEGER REFERENCES doctors(id),
    department_id INTEGER REFERENCES departments(id),
    notes         TEXT
);

-- ── set_items ────────────────────────────────
-- hospital_code: advisory reference only (not FK-enforced)
--   因套組耗材可能不在 items 主表中，故不設硬性 FK
-- quantity: NULL 表示 PRN（按需）
CREATE TABLE IF NOT EXISTS set_items (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    set_id        INTEGER NOT NULL REFERENCES sets(id) ON DELETE CASCADE,
    hospital_code TEXT,
    name_en       TEXT,
    name_zh       TEXT,
    quantity      INTEGER DEFAULT 1,
    price         INTEGER,
    notes         TEXT
);

-- ── indexes ──────────────────────────────────
CREATE INDEX IF NOT EXISTS idx_items_category   ON items(category);
CREATE INDEX IF NOT EXISTS idx_items_body_part  ON items(body_part);
CREATE INDEX IF NOT EXISTS idx_doctors_dept     ON doctors(department_id);
CREATE INDEX IF NOT EXISTS idx_sets_doctor      ON sets(doctor_id);
CREATE INDEX IF NOT EXISTS idx_sets_dept        ON sets(department_id);
CREATE INDEX IF NOT EXISTS idx_sets_surgery     ON sets(surgery_type);
CREATE INDEX IF NOT EXISTS idx_set_items_set    ON set_items(set_id);
CREATE INDEX IF NOT EXISTS idx_set_items_code   ON set_items(hospital_code);

"""

lines = [SCHEMA]

# INSERT doctors
lines.append('-- ── INSERT doctors ─────────────────────────\n')
for _, r in doctors_df.iterrows():
    lines.append(
        f"INSERT INTO doctors (id, name, department_id) VALUES "
        f"({sql_val(r['id'])}, {sql_val(r['name'])}, {sql_val(r['department_id'])});"
    )

# INSERT items
lines.append('\n-- ── INSERT items ───────────────────────────\n')
for _, r in items_df.iterrows():
    lines.append(
        f"INSERT INTO items (hospital_code, name_en, name_zh, category, body_part, unit, price, supplier, notes) VALUES "
        f"({sql_val(r['hospital_code'])}, {sql_val(r['name_en'])}, {sql_val(r['name_zh'])}, "
        f"{sql_val(r['category'])}, {sql_val(r['body_part'])}, {sql_val(r['unit'])}, "
        f"{sql_val(r['price'])}, {sql_val(r['supplier'])}, {sql_val(r['notes'])});"
    )

# INSERT sets
lines.append('\n-- ── INSERT sets ────────────────────────────\n')
for _, r in sets_df.iterrows():
    lines.append(
        f"INSERT INTO sets (id, name, surgery_type, doctor_id, department_id, notes) VALUES "
        f"({sql_val(r['id'])}, {sql_val(r['name'])}, {sql_val(r['surgery_type'])}, "
        f"{sql_val(r['doctor_id'])}, {sql_val(r['department_id'])}, {sql_val(r['notes'])});"
    )

# INSERT set_items
lines.append('\n-- ── INSERT set_items ───────────────────────\n')
for _, r in set_items_df.iterrows():
    lines.append(
        f"INSERT INTO set_items (id, set_id, hospital_code, name_en, name_zh, quantity, price, notes) VALUES "
        f"({sql_val(r['id'])}, {sql_val(r['set_id'])}, {sql_val(r['hospital_code'])}, "
        f"{sql_val(r['name_en'])}, {sql_val(r['name_zh'])}, {sql_val(r['quantity'])}, "
        f"{sql_val(r['price'])}, {sql_val(r['notes'])});"
    )

Path('data/medbase_schema.sql').write_text('\n'.join(lines), encoding='utf-8')
print("[OK] SQL  -> data/medbase_schema.sql")

# ═════════════════════════════════════════════
# OUTPUT 2 – Excel (5 sheets)
# ═════════════════════════════════════════════
HEADER_FILL = PatternFill('solid', start_color='4472C4', end_color='4472C4')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
CELL_FONT   = Font(name='Arial', size=10)
THIN_SIDE   = openpyxl.styles.borders.Side(style='thin', color='D9D9D9')
THIN_BORDER = openpyxl.styles.borders.Border(
    left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE
)
NOTE_FILL   = PatternFill('solid', start_color='FFF2CC', end_color='FFF2CC')

def write_sheet(wb, title, df, col_widths, notes=None):
    ws = wb.create_sheet(title)
    # header
    for c, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=c, value=col)
        cell.font   = HEADER_FONT
        cell.fill   = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(c)].width = col_widths.get(col, 14)
    ws.row_dimensions[1].height = 30
    # data
    for r, row in enumerate(df.itertuples(index=False), 2):
        for c, val in enumerate(row, 1):
            try:
                v = None if pd.isna(val) else val
            except (TypeError, ValueError):
                v = val
            cell = ws.cell(row=r, column=c, value=v)
            cell.font   = CELL_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical='center')
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    # optional note row at bottom
    if notes:
        nr = len(df) + 3
        nc = ws.cell(row=nr, column=1, value=notes)
        nc.font = Font(name='Arial', size=9, italic=True, color='595959')
        nc.fill = NOTE_FILL
    return ws

wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

# Sheet 1: items
write_sheet(wb, 'items', items_df, {
    'hospital_code': 14, 'name_en': 28, 'name_zh': 36,
    'category': 20, 'body_part': 16, 'unit': 8,
    'price': 10, 'supplier': 22, 'notes': 30,
}, notes='★ 以 hospital_code 為唯一主鍵；重複代碼已去除，保留最完整的一筆')

# Sheet 2: doctors
write_sheet(wb, 'doctors', doctors_df, {
    'id': 6, 'name': 36, 'department_id': 16,
}, notes='★ department_id 請填入 departments.id 完成科別關聯')

# Sheet 3: departments
write_sheet(wb, 'departments', departments_df, {
    'id': 6, 'name': 24,
}, notes='★ 此表為空，請手動填入科別（e.g., 骨科、胸腔外科）再回填 doctors.department_id')

# Sheet 4: sets
write_sheet(wb, 'sets', sets_df, {
    'id': 6, 'name': 40, 'surgery_type': 22,
    'doctor_id': 12, 'department_id': 16, 'notes': 30,
}, notes='★ doctor_id/department_id 兩者皆有=科別下特定醫師 | 僅doctor_id=個人常用 | 僅department_id=科別標準 | 皆無=術式通用')

# Sheet 5: set_items
write_sheet(wb, 'set_items', set_items_df, {
    'id': 6, 'set_id': 8, 'hospital_code': 14,
    'name_en': 26, 'name_zh': 32,
    'quantity': 8, 'price': 10, 'notes': 30,
}, notes='★ hospital_code 為 nullable FK；套組中的耗材若不在 items 主表中仍可保留')

wb.save('data/items_normalized.xlsx')
print("[OK] Excel -> data/items_normalized.xlsx")
