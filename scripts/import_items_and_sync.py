"""
匯入 items_purpose_dept.xlsx → 本地 SQLite + 合併更新雲端 (保留雲端專有項目)

使用方式：python scripts/import_items_and_sync.py
"""

import os, sys, json, sqlite3, urllib.request, urllib.error
import openpyxl

# ── 路徑設定 ──────────────────────────────────────────────────────
APPDATA    = os.environ["APPDATA"]
DB_PATH    = os.path.join(APPDATA, "com.medbase.app", "medbase.db")
XLSX_PATH  = os.path.join(os.path.dirname(__file__), "..", "data", "items_purpose_dept.xlsx")

# ── 1. 讀取 Excel ─────────────────────────────────────────────────
print("▶ 讀取 Excel…")
wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
ws = wb.active
headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
print("  欄位：", headers)

COL = {
    "hospital_code": None, "name_zh": None, "name_en": None,
    "purpose": None, "depts": None, "price": None, "supplier": None,
}
for i, h in enumerate(headers):
    if "hospital_code" in h: COL["hospital_code"] = i
    elif "name_zh"     in h: COL["name_zh"]      = i
    elif "name_en"     in h: COL["name_en"]      = i
    elif "purpose"     in h: COL["purpose"]      = i
    elif "depts"       in h: COL["depts"]        = i
    elif "price"       in h: COL["price"]        = i
    elif "supplier"    in h: COL["supplier"]     = i

missing = [k for k, v in COL.items() if v is None and k not in ("supplier",)]
if missing:
    sys.exit(f"❌ 找不到欄位：{missing}")

xlsx_rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    code = str(row[COL["hospital_code"]] or "").strip()
    if not code:
        continue
    name_zh  = str(row[COL["name_zh"]]  or "").strip()
    name_en  = str(row[COL["name_en"]]  or "").strip()
    purpose  = str(row[COL["purpose"]]  or "").strip()
    depts_raw= str(row[COL["depts"]]    or "").strip()
    price_v  = row[COL["price"]]
    supplier = str(row[COL["supplier"]] or "").strip() if COL["supplier"] is not None else ""

    price = None
    if price_v not in (None, ""):
        try: price = int(float(str(price_v)))
        except: pass

    depts = [d.strip() for d in depts_raw.split(";") if d.strip()]
    xlsx_rows.append({
        "hospital_code": code, "name_zh": name_zh, "name_en": name_en,
        "purpose": purpose, "price": price, "supplier": supplier, "depts": depts
    })

print(f"  Excel 有效列：{len(xlsx_rows)} 筆")

# ── 2. 寫入本地 SQLite ─────────────────────────────────────────────
print("▶ 連線本地資料庫…", DB_PATH)
con = sqlite3.connect(DB_PATH)
cur = con.cursor()

# 先確認 schema 存在
cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='items'")
if not cur.fetchone():
    sys.exit("❌ items 表不存在，請先啟動 app 初始化資料庫")

inserted = 0
updated  = 0
for r in xlsx_rows:
    cur.execute("SELECT hospital_code FROM items WHERE hospital_code=?", (r["hospital_code"],))
    exists = cur.fetchone()
    if exists:
        cur.execute("""
            UPDATE items SET name_zh=?, name_en=?, purpose=?, price=?, supplier=?
            WHERE hospital_code=?
        """, (r["name_zh"], r["name_en"], r["purpose"], r["price"], r["supplier"], r["hospital_code"]))
        updated += 1
    else:
        cur.execute("""
            INSERT INTO items (hospital_code, name_zh, name_en, purpose, price, supplier)
            VALUES (?,?,?,?,?,?)
        """, (r["hospital_code"], r["name_zh"], r["name_en"], r["purpose"], r["price"], r["supplier"]))
        inserted += 1

    # 科別：先刪後插（僅更新本次有資料的品項）
    cur.execute("DELETE FROM item_depts WHERE hospital_code=?", (r["hospital_code"],))
    for dept in r["depts"]:
        cur.execute("INSERT OR IGNORE INTO item_depts (hospital_code, dept) VALUES (?,?)", (r["hospital_code"], dept))

con.commit()
print(f"  新增：{inserted} 筆，更新：{updated} 筆")

# ── 3. 從資料庫讀取 GAS URL ───────────────────────────────────────
cur.execute("SELECT value FROM app_settings WHERE key='scheduler_gas_url'")
row = cur.fetchone()
if not row or not row[0]:
    con.close()
    print("⚠️  未設定 GAS URL，跳過雲端同步（資料已存入本地）")
    sys.exit(0)

GAS_URL = row[0].strip()
print(f"▶ GAS URL：{GAS_URL[:60]}…")

# ── 4. 拉取現有雲端 items ─────────────────────────────────────────
print("▶ 拉取雲端現有品項…")
try:
    req = urllib.request.Request(
        GAS_URL,
        data=json.dumps({"action": "getItems"}).encode(),
        headers={"Content-Type": "application/json"},
        method="POST"
    )
    with urllib.request.urlopen(req, timeout=30) as resp:
        cloud_resp = json.loads(resp.read())
except Exception as e:
    con.close()
    sys.exit(f"❌ 拉取雲端失敗：{e}")

cloud_items = cloud_resp.get("data", [])
print(f"  雲端現有：{len(cloud_items)} 筆")

# ── 5. 合併：import 清單覆蓋（直接用 xlsx 資料），雲端專有保持不變 ──
import_codes = {r["hospital_code"] for r in xlsx_rows}

cloud_only = [item for item in cloud_items if item["hospital_code"] not in import_codes]
print(f"  雲端專有（保留不動）：{len(cloud_only)} 筆")

# xlsx 品項：直接以 xlsx 欄位推送，不繞道 DB（避免 None 遮蔽問題）
push_list = []
for r in xlsx_rows:
    push_list.append({
        "hospital_code": r["hospital_code"],
        "name_zh":  r["name_zh"],
        "name_en":  r["name_en"],
        "purpose":  r["purpose"],
        "unit":     "",
        "price":    r["price"],
        "supplier": r["supplier"],
        "notes":    "",
        "depts":    r["depts"],
    })

# 雲端專有品項原樣附加
for item in cloud_only:
    push_list.append(item)

print(f"  合併後推送總筆數：{len(push_list)}")

# ── 6. 推送到雲端 ─────────────────────────────────────────────────
print("▶ 推送至雲端…")
try:
    req2 = urllib.request.Request(
        GAS_URL,
        data=json.dumps({"action": "saveItems", "data": push_list}).encode(),
        headers={"Content-Type": "application/json"},
        method="POST"
    )
    with urllib.request.urlopen(req2, timeout=60) as resp2:
        save_resp = json.loads(resp2.read())
except Exception as e:
    con.close()
    sys.exit(f"❌ 推送雲端失敗：{e}")

con.close()

if save_resp.get("ok"):
    print(f"✅ 完成！本地新增 {inserted} / 更新 {updated}，雲端保留 {len(cloud_only)} 筆專有品項，共 {len(push_list)} 筆")
else:
    print(f"⚠️  雲端回應異常：{save_resp}")
